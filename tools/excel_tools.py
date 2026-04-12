"""
Excel Tools - работа с таблицами
"""
import os
from pathlib import Path


class ExcelTools:

    def read_excel(self, path: str, sheet: str = None) -> str:
        """Читать Excel файл"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb[sheet] if sheet else wb.active
            result = []
            for row in ws.iter_rows(values_only=True):
                row_data = [str(c) if c is not None else "" for c in row]
                if any(row_data):
                    result.append(" | ".join(row_data))
            return f"Sheet: {ws.title}\n" + "\n".join(result[:50])
        except ImportError:
            return "Install openpyxl: pip install openpyxl"
        except Exception as e:
            return f"Error: {e}"

    def write_excel(self, path: str, data: list, headers: list = None) -> str:
        """Создать Excel файл"""
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            if headers:
                ws.append(headers)
            for row in data:
                ws.append(row if isinstance(row, list) else [row])
            wb.save(path)
            return f"✅ Excel saved: {path}"
        except ImportError:
            return "Install openpyxl: pip install openpyxl"
        except Exception as e:
            return f"Error: {e}"

    def create_table(self, path: str, text_data: str) -> str:
        """Создать таблицу из текстовых данных"""
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            lines = [l.strip() for l in text_data.strip().split('\n') if l.strip()]
            for line in lines:
                if ',' in line:
                    row = [c.strip() for c in line.split(',')]
                elif '|' in line:
                    row = [c.strip() for c in line.split('|') if c.strip()]
                elif '\t' in line:
                    row = [c.strip() for c in line.split('\t')]
                else:
                    row = [line]
                ws.append(row)
            wb.save(path)
            return f"✅ Table created: {path} ({len(lines)} rows)"
        except Exception as e:
            return f"Error: {e}"

    def add_row(self, path: str, row_data: list) -> str:
        """Добавить строку в Excel"""
        try:
            import openpyxl
            if Path(path).exists():
                wb = openpyxl.load_workbook(path)
            else:
                wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(row_data)
            wb.save(path)
            return f"✅ Row added to {path}"
        except Exception as e:
            return f"Error: {e}"

    def get_sheets(self, path: str) -> str:
        """Список листов в Excel"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path)
            return f"Sheets: {', '.join(wb.sheetnames)}"
        except Exception as e:
            return f"Error: {e}"

    def read_csv(self, path: str) -> str:
        """Читать CSV файл"""
        try:
            import csv
            rows = []
            with open(path, newline='', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                for i, row in enumerate(reader):
                    rows.append(" | ".join(row))
                    if i > 50:
                        rows.append("... (truncated)")
                        break
            return "\n".join(rows)
        except Exception as e:
            return f"Error: {e}"

    def write_csv(self, path: str, data: str) -> str:
        """Создать CSV файл"""
        try:
            import csv
            lines = [l.strip() for l in data.strip().split('\n') if l.strip()]
            with open(path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                for line in lines:
                    writer.writerow([c.strip() for c in line.split(',')])
            return f"✅ CSV saved: {path}"
        except Exception as e:
            return f"Error: {e}"
